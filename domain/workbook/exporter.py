from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import WorkbookModel

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
DEFAULT_ALIGNMENT = Alignment(horizontal="left", vertical="center")


def export_workbook_to_xlsx(workbook: WorkbookModel, output_path: Path) -> Path:
    workbook_writer = Workbook()
    default_sheet = workbook_writer.active

    if workbook.worksheets:
        workbook_writer.remove(default_sheet)
        for worksheet_model in workbook.worksheets:
            worksheet = workbook_writer.create_sheet(title=worksheet_model.name)
            _write_worksheet_rows(worksheet, worksheet_model.rows)
    else:
        default_sheet.title = "Sheet1"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook_writer.save(output_path)
    return output_path


def _write_worksheet_rows(worksheet, rows: list[list[object]]) -> None:
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = DEFAULT_ALIGNMENT
            if row_index == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL


__all__ = ["export_workbook_to_xlsx"]
