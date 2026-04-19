from datetime import datetime, timedelta, timezone
from pathlib import Path
from threading import Event, Thread

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from astrbot_plugin_office_assistant.domain.workbook.contracts import (
    CreateWorkbookRequest,
    ExportWorkbookRequest,
    MAX_WORKBOOK_ROW_INDEX,
    WriteRowsRequest,
    build_workbook_summary,
)
from astrbot_plugin_office_assistant.domain.workbook.session_store import (
    WorkbookSessionStore,
)
import astrbot_plugin_office_assistant.domain.workbook.session_store as workbook_session_store_module


def test_create_workbook_returns_draft_summary(workspace_root: Path):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(
        CreateWorkbookRequest(
            session_id="pytest-session",
            title="季度销量汇总",
            filename="sales-report",
        )
    )

    summary = store.build_prompt_summary(workbook.workbook_id)

    assert workbook.workbook_id == "wb-1"
    assert workbook.metadata.title == "季度销量汇总"
    assert workbook.metadata.preferred_filename == "sales-report.xlsx"
    assert summary == {
        "workbook_id": workbook.workbook_id,
        "title": "季度销量汇总",
        "status": "draft",
        "sheet_names": [],
        "sheet_count": 0,
        "latest_written_sheets": [],
        "next_allowed_actions": ["write_rows", "export_workbook"],
    }


def test_workbook_session_store_evicts_oldest_workbooks_when_capped():
    store = WorkbookSessionStore(max_workbooks=2)
    first = store.create_workbook(CreateWorkbookRequest(title="first"))
    second = store.create_workbook(CreateWorkbookRequest(title="second"))
    third = store.create_workbook(CreateWorkbookRequest(title="third"))

    assert store.get_workbook(first.workbook_id) is None
    assert store.get_workbook(second.workbook_id) is not None
    assert store.get_workbook(third.workbook_id) is not None


def test_workbook_session_store_evicts_expired_workbooks_by_ttl():
    store = WorkbookSessionStore(ttl=timedelta(seconds=1))
    expired = store.create_workbook(CreateWorkbookRequest(title="expired"))
    fresh = store.create_workbook(CreateWorkbookRequest(title="fresh"))

    expired.metadata.updated_at = datetime.now(timezone.utc) - timedelta(seconds=5)
    fresh.metadata.updated_at = datetime.now(timezone.utc)

    assert store.get_workbook(expired.workbook_id) is None
    assert store.get_workbook(fresh.workbook_id) is not None


def test_write_rows_creates_sheet_and_overwrites_range(workspace_root: Path):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="rows.xlsx"))

    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Summary",
            rows=[
                ["name", "amount"],
                ["A", 10],
            ],
            start_row=1,
        )
    )
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Summary",
            rows=[
                ["A", 99],
                ["B", 35],
            ],
            start_row=2,
        )
    )

    loaded = store.require_workbook(workbook.workbook_id)
    worksheet = loaded.get_sheet("Summary")
    assert worksheet is not None
    assert worksheet.rows == [
        ["name", "amount"],
        ["A", 99],
        ["B", 35],
    ]

    prompt_summary = store.build_prompt_summary(workbook.workbook_id)
    assert prompt_summary["sheet_names"] == ["Summary"]
    assert prompt_summary["latest_written_sheets"] == ["Summary"]


def test_write_rows_reuses_sheet_with_case_insensitive_name(workspace_root: Path):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="rows.xlsx"))

    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Sales",
            rows=[["name", "amount"]],
            start_row=1,
        )
    )
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="sales",
            rows=[["A", 10]],
            start_row=2,
        )
    )

    loaded = store.require_workbook(workbook.workbook_id)
    assert len(loaded.worksheets) == 1
    assert loaded.worksheets[0].name == "Sales"
    assert loaded.worksheets[0].rows == [["name", "amount"], ["A", 10]]

    prompt_summary = store.build_prompt_summary(workbook.workbook_id)
    assert prompt_summary["sheet_names"] == ["Sales"]
    assert prompt_summary["latest_written_sheets"] == ["Sales"]


def test_write_rows_rejects_non_primitive_cell_values():
    with pytest.raises(ValidationError):
        WriteRowsRequest(
            workbook_id="wb-1",
            sheet="Data",
            rows=[[{"invalid": True}]],  # type: ignore[list-item]
            start_row=1,
        )


def test_write_rows_rejects_formula_cells():
    with pytest.raises(ValidationError, match="not supported"):
        WriteRowsRequest(
            workbook_id="wb-1",
            sheet="Data",
            rows=[["=SUM(A1:A2)"]],
            start_row=1,
        )


def test_write_rows_rejects_start_row_beyond_limit():
    with pytest.raises(ValidationError, match="less than or equal to"):
        WriteRowsRequest(
            workbook_id="wb-1",
            sheet="Data",
            rows=[["value"]],
            start_row=MAX_WORKBOOK_ROW_INDEX + 1,
        )


def test_write_rows_rejects_final_row_beyond_limit():
    with pytest.raises(ValidationError, match="final written row must not exceed"):
        WriteRowsRequest(
            workbook_id="wb-1",
            sheet="Data",
            rows=[["value"], ["overflow"]],
            start_row=MAX_WORKBOOK_ROW_INDEX,
        )


@pytest.mark.parametrize(
    "output_name",
    [
        "C:/temp/final.xlsx",
        "/tmp/final.xlsx",
        "~/final.xlsx",
    ],
)
def test_export_request_rejects_absolute_output_name(output_name: str):
    with pytest.raises(ValidationError, match="must not be an absolute path"):
        ExportWorkbookRequest(
            workbook_id="wb-1",
            output_name=output_name,
        )


def test_export_rejects_output_path_outside_workspace(workspace_root: Path):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="status.xlsx"))
    escaped_request = ExportWorkbookRequest.model_construct(
        workbook_id=workbook.workbook_id,
        output_name="../evil.xlsx",
    )

    with pytest.raises(ValueError, match="escape the workbook workspace"):
        store.export_workbook(escaped_request)


def test_exporter_writes_values_and_header_style(workspace_root: Path):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="styled.xlsx"))
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Sales",
            rows=[
                ["name", "amount"],
                ["A", 123.5],
            ],
        )
    )

    exported_workbook, output_path = store.export_workbook(
        ExportWorkbookRequest(
            workbook_id=workbook.workbook_id,
            output_name="styled-output.xlsx",
        )
    )

    loaded = load_workbook(output_path)
    sheet = loaded["Sales"]

    assert exported_workbook.status.value == "exported"
    assert output_path.name == "styled-output.xlsx"
    assert sheet["A1"].value == "name"
    assert sheet["B2"].value == 123.5
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].fill.patternType == "solid"
    assert (sheet["A1"].fill.fgColor.rgb or "").endswith("D9D9D9")
    assert sheet["A1"].alignment.horizontal == "left"
    assert sheet["A1"].alignment.vertical == "center"
    assert sheet["A2"].alignment.horizontal == "left"
    assert sheet["A2"].font.bold is False


def test_exporter_applies_worksheet_options(workspace_root: Path):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="options.xlsx"))
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Sales",
            rows=[
                ["name", "amount"],
                ["A", 123.5],
            ],
        )
    )
    worksheet = store.require_workbook(workbook.workbook_id).get_sheet("Sales")
    assert worksheet is not None
    worksheet.options.freeze_panes = "B2"
    worksheet.options.column_widths = {"A": 20.0, "B": 15.5}
    worksheet.options.autofilter = True

    _, output_path = store.export_workbook(
        ExportWorkbookRequest(
            workbook_id=workbook.workbook_id,
            output_name="options-output.xlsx",
        )
    )

    loaded = load_workbook(output_path)
    sheet = loaded["Sales"]

    assert sheet.freeze_panes == "B2"
    assert sheet.column_dimensions["A"].width == pytest.approx(20.0)
    assert sheet.column_dimensions["B"].width == pytest.approx(15.5)
    assert sheet.auto_filter.ref == "A1:B2"


def test_exported_workbook_disallows_more_writes_or_reexport(workspace_root: Path):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="status.xlsx"))
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Data",
            rows=[["h1"], ["v1"]],
        )
    )
    store.export_workbook(ExportWorkbookRequest(workbook_id=workbook.workbook_id))

    with pytest.raises(ValueError, match="status is draft"):
        store.write_rows(
            WriteRowsRequest(
                workbook_id=workbook.workbook_id,
                sheet="Data",
                rows=[["v2"]],
                start_row=3,
            )
        )

    with pytest.raises(ValueError, match="status is draft"):
        store.export_workbook(
            ExportWorkbookRequest(workbook_id=workbook.workbook_id)
        )

    summary = store.build_prompt_summary(workbook.workbook_id)
    assert summary["status"] == "exported"
    assert summary["next_allowed_actions"] == []


def test_exported_workbook_compacts_rows_but_keeps_summary(workspace_root: Path):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="compact.xlsx"))
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Data",
            rows=[["h1"], ["v1"]],
        )
    )

    store.export_workbook(ExportWorkbookRequest(workbook_id=workbook.workbook_id))

    exported_workbook = store.require_workbook(workbook.workbook_id)
    assert exported_workbook.worksheets[0].name == "Data"
    assert exported_workbook.worksheets[0].rows == []

    summary = store.build_prompt_summary(workbook.workbook_id)
    assert summary["status"] == "exported"
    assert summary["sheet_names"] == ["Data"]
    assert summary["sheet_count"] == 1


def test_export_workbook_rejects_same_workbook_writes_without_waiting_for_export(
    workspace_root: Path, monkeypatch: pytest.MonkeyPatch
):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="status.xlsx"))
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Data",
            rows=[["h1"], ["v1"]],
        )
    )

    export_started = Event()
    release_export = Event()
    write_finished = Event()
    write_error: list[Exception] = []

    def fake_export(workbook_model, output_path):
        export_started.set()
        assert not write_finished.is_set()
        release_export.wait(timeout=5)
        output_path.write_text("xlsx", encoding="utf-8")
        return output_path

    def run_export():
        store.export_workbook(ExportWorkbookRequest(workbook_id=workbook.workbook_id))

    def run_write():
        try:
            store.write_rows(
                WriteRowsRequest(
                    workbook_id=workbook.workbook_id,
                    sheet="Data",
                    rows=[["v2"]],
                    start_row=3,
                )
            )
        except Exception as exc:  # pragma: no cover - asserted via captured error
            write_error.append(exc)
        finally:
            write_finished.set()

    monkeypatch.setattr(workbook_session_store_module, "export_workbook_to_xlsx", fake_export)

    export_thread = Thread(target=run_export)
    write_thread = Thread(target=run_write)

    export_thread.start()
    assert export_started.wait(timeout=5)
    write_thread.start()

    assert write_finished.wait(timeout=0.2)

    release_export.set()
    export_thread.join(timeout=5)
    write_thread.join(timeout=5)

    assert len(write_error) == 1
    assert "status is draft" in str(write_error[0])


def test_export_workbook_allows_other_workbook_operations_while_writing_file(
    workspace_root: Path, monkeypatch: pytest.MonkeyPatch
):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="status.xlsx"))
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Data",
            rows=[["h1"], ["v1"]],
        )
    )

    export_started = Event()
    release_export = Event()
    other_finished = Event()
    other_error: list[Exception] = []
    other_workbook_ids: list[str] = []

    def fake_export(workbook_model, output_path):
        export_started.set()
        release_export.wait(timeout=5)
        output_path.write_text("xlsx", encoding="utf-8")
        return output_path

    def run_export():
        store.export_workbook(ExportWorkbookRequest(workbook_id=workbook.workbook_id))

    def run_other_workbook_ops():
        try:
            other = store.create_workbook(CreateWorkbookRequest(filename="other.xlsx"))
            other_workbook_ids.append(other.workbook_id)
            store.write_rows(
                WriteRowsRequest(
                    workbook_id=other.workbook_id,
                    sheet="Data",
                    rows=[["value"]],
                )
            )
        except Exception as exc:  # pragma: no cover - asserted via captured error
            other_error.append(exc)
        finally:
            other_finished.set()

    monkeypatch.setattr(workbook_session_store_module, "export_workbook_to_xlsx", fake_export)

    export_thread = Thread(target=run_export)
    other_thread = Thread(target=run_other_workbook_ops)

    export_thread.start()
    assert export_started.wait(timeout=5)
    assert store.build_prompt_summary(workbook.workbook_id)["status"] == "exporting"

    other_thread.start()
    assert other_finished.wait(timeout=0.2)

    release_export.set()
    export_thread.join(timeout=5)
    other_thread.join(timeout=5)

    assert other_error == []
    assert other_workbook_ids == ["wb-2"]
    other_summary = store.build_prompt_summary(other_workbook_ids[0])
    assert other_summary["status"] == "draft"
    assert other_summary["next_allowed_actions"] == ["write_rows", "export_workbook"]


def test_create_workbook_keeps_new_draft_when_store_is_capped_by_exporting_workbook(
    workspace_root: Path, monkeypatch: pytest.MonkeyPatch
):
    store = WorkbookSessionStore(workspace_dir=workspace_root, max_workbooks=1)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="status.xlsx"))
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Data",
            rows=[["h1"], ["v1"]],
        )
    )

    export_started = Event()
    release_export = Event()
    create_finished = Event()
    create_error: list[Exception] = []
    created_workbook_ids: list[str] = []

    def fake_export(workbook_model, output_path):
        export_started.set()
        release_export.wait(timeout=5)
        output_path.write_text("xlsx", encoding="utf-8")
        return output_path

    def run_export():
        store.export_workbook(ExportWorkbookRequest(workbook_id=workbook.workbook_id))

    def run_create_and_write():
        try:
            other = store.create_workbook(CreateWorkbookRequest(filename="other.xlsx"))
            created_workbook_ids.append(other.workbook_id)
            store.write_rows(
                WriteRowsRequest(
                    workbook_id=other.workbook_id,
                    sheet="Data",
                    rows=[["value"]],
                )
            )
        except Exception as exc:  # pragma: no cover - asserted via captured error
            create_error.append(exc)
        finally:
            create_finished.set()

    monkeypatch.setattr(workbook_session_store_module, "export_workbook_to_xlsx", fake_export)

    export_thread = Thread(target=run_export)
    create_thread = Thread(target=run_create_and_write)

    export_thread.start()
    assert export_started.wait(timeout=5)

    create_thread.start()
    assert create_finished.wait(timeout=0.2)

    assert create_error == []
    assert created_workbook_ids == ["wb-2"]
    created_summary = store.build_prompt_summary(created_workbook_ids[0])
    assert created_summary["status"] == "draft"
    assert created_summary["next_allowed_actions"] == ["write_rows", "export_workbook"]

    release_export.set()
    export_thread.join(timeout=5)
    create_thread.join(timeout=5)

    assert store.get_workbook(workbook.workbook_id) is None
    assert store.get_workbook(created_workbook_ids[0]) is not None
    created_summary = store.build_prompt_summary(created_workbook_ids[0])
    assert created_summary["status"] == "draft"
    assert created_summary["next_allowed_actions"] == ["write_rows", "export_workbook"]


def test_export_workbook_resets_status_after_export_failure(
    workspace_root: Path, monkeypatch: pytest.MonkeyPatch
):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="status.xlsx"))
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Data",
            rows=[["h1"], ["v1"]],
        )
    )

    def fake_export(_workbook_model, _output_path):
        raise RuntimeError("disk full")

    monkeypatch.setattr(workbook_session_store_module, "export_workbook_to_xlsx", fake_export)

    with pytest.raises(RuntimeError, match="disk full"):
        store.export_workbook(ExportWorkbookRequest(workbook_id=workbook.workbook_id))

    summary = store.build_prompt_summary(workbook.workbook_id)
    assert summary["status"] == "draft"
    assert summary["next_allowed_actions"] == ["write_rows", "export_workbook"]

    loaded = store.require_workbook(workbook.workbook_id)
    assert loaded.output_path == ""


def test_build_workbook_summary_includes_sheet_names(workspace_root: Path):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="book.xlsx"))
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="One",
            rows=[["c1"], ["v1"]],
        )
    )
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Two",
            rows=[["c1"], ["v2"]],
        )
    )

    summary = build_workbook_summary(store.require_workbook(workbook.workbook_id))

    assert summary.workbook_id == workbook.workbook_id
    assert summary.sheet_count == 2
    assert summary.sheet_names == ["One", "Two"]
    assert summary.latest_written_sheets == ["One", "Two"]


def test_prompt_summary_reuses_shared_workbook_summary_fields(workspace_root: Path):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(
        CreateWorkbookRequest(
            session_id="pytest-session",
            title="季度汇总",
            filename="book.xlsx",
        )
    )
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="总表",
            rows=[["列1"], ["值1"]],
        )
    )

    workbook_summary = build_workbook_summary(
        store.require_workbook(workbook.workbook_id)
    ).model_dump()
    prompt_summary = store.build_prompt_summary(workbook.workbook_id)

    shared_keys = (
        "workbook_id",
        "title",
        "status",
        "sheet_names",
        "sheet_count",
        "latest_written_sheets",
    )
    assert {key: prompt_summary[key] for key in shared_keys} == {
        key: workbook_summary[key] for key in shared_keys
    }


def test_case_insensitive_sheet_lookup_preserves_exported_sheet_name(workspace_root: Path):
    store = WorkbookSessionStore(workspace_dir=workspace_root)
    workbook = store.create_workbook(CreateWorkbookRequest(filename="book.xlsx"))
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="Sales",
            rows=[["name"], ["A"]],
        )
    )
    store.write_rows(
        WriteRowsRequest(
            workbook_id=workbook.workbook_id,
            sheet="sales",
            rows=[["B"]],
            start_row=3,
        )
    )

    _, output_path = store.export_workbook(
        ExportWorkbookRequest(workbook_id=workbook.workbook_id)
    )
    loaded = load_workbook(output_path)

    assert loaded.sheetnames == ["Sales"]
    sheet = loaded["Sales"]
    assert sheet["A3"].value == "B"
