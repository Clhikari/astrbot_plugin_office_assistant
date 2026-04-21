from __future__ import annotations

import asyncio
import inspect
import json
import shutil
import subprocess
import sys
import tempfile
import textwrap
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from uuid import uuid4

from astrbot.api import logger
from astrbot.api.event import AstrMessageEvent

from ..constants import EXCEL_SUFFIXES

try:
    from astrbot.core.computer.computer_client import (
        get_booter as _get_computer_booter,
    )
except ImportError:  # pragma: no cover - depends on AstrBot runtime
    _get_computer_booter = None


@dataclass(frozen=True, slots=True)
class ScriptProcessResult:
    success: bool
    mode: str
    result_text: str | None = None
    output_path: Path | None = None
    error: str | None = None
    traceback: str | None = None
    script: str | None = None
    retryable: bool = True


@dataclass(frozen=True, slots=True)
class SandboxScriptPaths:
    exec_dir: str
    result_path: str
    remote_result_path: str
    input_files: list[str]
    remote_input_files: list[str]
    output_path: str | None = None
    remote_output_path: str | None = None


class ExcelScriptService:
    _EXCEL_SUFFIXES = EXCEL_SUFFIXES
    _MAX_SCRIPT_RETRIES = 2
    _RETRY_COUNT_EVENT_KEY = "office_assistant_excel_script_retry_failures"
    _SCRIPT_TIMEOUT_SECONDS = 30
    _SANDBOX_WORKSPACE_ROOT = "/workspace"
    _SANDBOX_EXEC_ROOT = PurePosixPath(".office_assistant") / "excel_scripts"

    def __init__(
        self,
        *,
        astrbot_context=None,
        workspace_service,
        file_delivery_service,
        allow_external_input_files: bool,
        is_group_feature_enabled,
        check_permission,
        group_feature_disabled_error,
    ) -> None:
        self._astrbot_context = astrbot_context
        self._workspace_service = workspace_service
        self._file_delivery_service = file_delivery_service
        self._allow_external_input_files = allow_external_input_files
        self._is_group_feature_enabled = is_group_feature_enabled
        self._check_permission = check_permission
        self._group_feature_disabled_error = group_feature_disabled_error

    @staticmethod
    def _build_error_result(
        *,
        script: str,
        error: str,
        traceback_text: str = "",
        retry_count: int = 0,
        retry_exhausted: bool = False,
    ) -> str:
        return json.dumps(
            {
                "success": False,
                "error": error,
                "traceback": traceback_text,
                "script": script,
                "retry_count": retry_count,
                "max_retries": ExcelScriptService._MAX_SCRIPT_RETRIES,
                "retry_exhausted": retry_exhausted,
            },
            ensure_ascii=False,
        )

    @classmethod
    def _get_failure_count(cls, event: AstrMessageEvent) -> int:
        get_extra = getattr(event, "get_extra", None)
        if not callable(get_extra):
            return 0
        value = get_extra(cls._RETRY_COUNT_EVENT_KEY, 0)
        try:
            return max(int(value), 0)
        except (TypeError, ValueError):
            return 0

    @classmethod
    def _set_failure_count(cls, event: AstrMessageEvent, count: int) -> None:
        set_extra = getattr(event, "set_extra", None)
        if callable(set_extra):
            set_extra(cls._RETRY_COUNT_EVENT_KEY, max(count, 0))

    @classmethod
    def _reset_failure_count(cls, event: AstrMessageEvent) -> None:
        cls._set_failure_count(event, 0)

    @classmethod
    def _retry_used_count(cls, failure_count: int) -> int:
        return max(failure_count - 1, 0)

    @classmethod
    def _build_retry_exhausted_error(cls, error: str) -> str:
        return (
            f"{error}；已超过最多 {cls._MAX_SCRIPT_RETRIES} 次脚本重试，请停止重试，"
            "直接向用户说明失败原因，并建议缩小需求或改走原语路径。"
        )

    @classmethod
    def _build_exhausted_result(
        cls,
        *,
        script: str,
        error: str,
        traceback_text: str = "",
        failure_count: int,
    ) -> str:
        return cls._build_error_result(
            script=script,
            error=cls._build_retry_exhausted_error(error),
            traceback_text=traceback_text,
            retry_count=cls._retry_used_count(failure_count),
            retry_exhausted=True,
        )

    def _build_failure_result(
        self,
        event: AstrMessageEvent,
        *,
        script: str,
        error: str,
        traceback_text: str = "",
    ) -> str:
        failure_count = self._get_failure_count(event) + 1
        self._set_failure_count(event, failure_count)
        if failure_count > self._MAX_SCRIPT_RETRIES:
            return self._build_exhausted_result(
                script=script,
                error=error,
                traceback_text=traceback_text,
                failure_count=failure_count,
            )
        return self._build_error_result(
            script=script,
            error=error,
            traceback_text=traceback_text,
            retry_count=self._retry_used_count(failure_count),
            retry_exhausted=False,
        )

    def _build_non_retry_result(
        self,
        event: AstrMessageEvent,
        *,
        script: str,
        error: str,
        traceback_text: str = "",
    ) -> str:
        failure_count = self._get_failure_count(event)
        return self._build_error_result(
            script=script,
            error=error,
            traceback_text=traceback_text,
            retry_count=self._retry_used_count(failure_count),
            retry_exhausted=False,
        )

    async def execute_excel_script(
        self,
        event: AstrMessageEvent,
        *,
        script: str,
        input_files: list[str] | None = None,
        output_name: str | None = None,
    ) -> str:
        normalized_script = (script or "").strip()
        current_failure_count = self._get_failure_count(event)
        if current_failure_count > self._MAX_SCRIPT_RETRIES:
            return self._build_exhausted_result(
                script=normalized_script or script or "",
                error="脚本重试次数已用尽",
                failure_count=current_failure_count,
            )
        if not normalized_script:
            return self._build_non_retry_result(
                event,
                script=script or "",
                error="缺少 script 参数",
            )

        if not self._is_group_feature_enabled(event):
            return self._build_non_retry_result(
                event,
                script=normalized_script,
                error=self._group_feature_disabled_error(),
            )

        if not self._check_permission(event):
            return self._build_non_retry_result(
                event,
                script=normalized_script,
                error="错误：权限不足",
            )

        resolved_input_files: list[Path] = []
        for filename in input_files or []:
            ok, resolved_path, err = self._workspace_service.pre_check(
                event,
                filename,
                require_exists=True,
                allowed_suffixes=self._EXCEL_SUFFIXES,
                allow_external_path=self._allow_external_input_files,
                is_group_feature_enabled=self._is_group_feature_enabled,
                check_permission_fn=self._check_permission,
                group_feature_disabled_error=self._group_feature_disabled_error,
            )
            if not ok:
                return self._build_non_retry_result(
                    event,
                    script=normalized_script,
                    error=err or "错误：未知错误",
                )
            if resolved_path is None:
                return self._build_non_retry_result(
                    event,
                    script=normalized_script,
                    error="错误：输入文件路径解析失败",
                )
            resolved_input_files.append(resolved_path)

        resolved_output_path: Path | None = None
        if output_name:
            ok, resolved_path, err = self._workspace_service.pre_check(
                event,
                output_name,
                require_exists=False,
                allowed_suffixes=self._EXCEL_SUFFIXES,
                allow_external_path=False,
                is_group_feature_enabled=self._is_group_feature_enabled,
                check_permission_fn=self._check_permission,
                group_feature_disabled_error=self._group_feature_disabled_error,
            )
            if not ok:
                return self._build_non_retry_result(
                    event,
                    script=normalized_script,
                    error=err or "错误：未知错误",
                )
            if resolved_path is None:
                return self._build_non_retry_result(
                    event,
                    script=normalized_script,
                    error="错误：输出文件路径解析失败",
                )
            resolved_output_path = resolved_path
            resolved_output_path.parent.mkdir(parents=True, exist_ok=True)

        runtime_mode = self._resolve_runtime_mode(event)
        if runtime_mode == "none":
            return self._build_non_retry_result(
                event,
                script=normalized_script,
                error=(
                    "错误：当前 computer runtime 为 none，无法执行 Excel 脚本，"
                    "请启用 local 或 sandbox，或改走原语路径"
                ),
            )

        booter = None
        if runtime_mode == "sandbox":
            booter, booter_error = await self._acquire_sandbox_booter(event)
            if booter is None:
                return self._build_non_retry_result(
                    event,
                    script=normalized_script,
                    error=booter_error or "错误：Excel sandbox 不可用",
                )

        process_result = await self._run_script_process(
            script=normalized_script,
            input_files=resolved_input_files,
            output_path=resolved_output_path,
            runtime_mode=runtime_mode,
            booter=booter,
        )

        if not process_result.success:
            build_error_result = (
                self._build_failure_result
                if getattr(process_result, "retryable", True)
                else self._build_non_retry_result
            )
            return build_error_result(
                event,
                script=process_result.script or normalized_script,
                error=process_result.error or "脚本执行失败",
                traceback_text=process_result.traceback or "",
            )

        if process_result.mode == "text":
            self._reset_failure_count(event)
            return json.dumps(
                {
                    "success": True,
                    "mode": "text",
                    "result_text": process_result.result_text or "",
                },
                ensure_ascii=False,
            )

        delivery_error = await self._file_delivery_service.deliver_generated_file(
            event,
            process_result.output_path,
            missing_message="错误：Excel 脚本执行成功，但未找到生成的文件",
            oversized_template=(
                "错误：生成的 Excel 文件大小 {file_size} 超过限制 {max_size}"
            ),
        )
        if delivery_error:
            return self._build_failure_result(
                event,
                script=normalized_script,
                error=delivery_error,
            )

        self._reset_failure_count(event)
        return json.dumps(
            {
                "success": True,
                "mode": "file",
                "output_name": process_result.output_path.name
                if process_result.output_path is not None
                else "",
            },
            ensure_ascii=False,
        )

    def _resolve_runtime_mode(self, event: AstrMessageEvent) -> str:
        if self._astrbot_context is None:
            return "local"
        get_config = getattr(self._astrbot_context, "get_config", None)
        if not callable(get_config):
            return "local"
        session_id = str(getattr(event, "unified_msg_origin", "") or "")
        config = self._get_session_config(get_config, session_id)
        if not isinstance(config, dict):
            return "local"
        provider_settings = config.get("provider_settings", {})
        if not isinstance(provider_settings, dict):
            return "local"
        runtime = provider_settings.get("computer_use_runtime", "local")
        if isinstance(runtime, str) and runtime.strip():
            return runtime.strip().lower()
        return "local"

    @staticmethod
    def _get_session_config(get_config, session_id: str):
        try:
            signature = inspect.signature(get_config)
        except (TypeError, ValueError):
            try:
                return get_config(session_id)
            except TypeError:
                try:
                    return get_config(umo=session_id)
                except TypeError:
                    return get_config()

        parameters = tuple(signature.parameters.values())
        if any(
            parameter.kind is inspect.Parameter.VAR_KEYWORD for parameter in parameters
        ) or "umo" in signature.parameters:
            return get_config(umo=session_id)

        if any(
            parameter.kind
            in (
                inspect.Parameter.POSITIONAL_ONLY,
                inspect.Parameter.POSITIONAL_OR_KEYWORD,
                inspect.Parameter.VAR_POSITIONAL,
            )
            for parameter in parameters
        ):
            return get_config(session_id)

        return get_config()

    async def _acquire_sandbox_booter(
        self,
        event: AstrMessageEvent,
    ) -> tuple[object | None, str | None]:
        if self._astrbot_context is None:
            return None, "错误：当前服务未注入 AstrBot context，无法使用 sandbox"
        if _get_computer_booter is None:
            return None, "错误：当前 AstrBot 版本未提供 sandbox 执行接口"
        if self._resolve_runtime_mode(event) != "sandbox":
            return (
                None,
                "错误：execute_excel_script 仅支持 sandbox runtime，请在 AstrBot 中启用 sandbox",
            )
        session_id = str(getattr(event, "unified_msg_origin", "") or "").strip()
        if not session_id:
            return None, "错误：缺少会话标识，无法初始化 Excel sandbox"
        try:
            booter = await _get_computer_booter(self._astrbot_context, session_id)
        except Exception as exc:
            logger.warning(f"[文件管理] Excel sandbox 初始化失败: {exc}")
            return None, f"错误：Excel sandbox 初始化失败：{exc}"

        capabilities = getattr(booter, "capabilities", None)
        if capabilities is not None:
            required_capabilities = {"python", "filesystem"}
            missing = sorted(required_capabilities.difference(capabilities))
            if missing:
                missing_text = ", ".join(missing)
                return (
                    None,
                    f"错误：当前 sandbox profile 缺少必要能力：{missing_text}",
                )
        return booter, None

    @classmethod
    def _join_sandbox_path(cls, workspace_root: str, relative_path: PurePosixPath) -> str:
        if len(workspace_root) >= 2 and workspace_root[1] == ":":
            return str(Path(workspace_root) / Path(relative_path.as_posix()))
        if "\\" in workspace_root:
            return str(Path(workspace_root) / Path(relative_path.as_posix()))
        return str(PurePosixPath(workspace_root or cls._SANDBOX_WORKSPACE_ROOT) / relative_path)

    @classmethod
    def _build_sandbox_paths(
        cls,
        booter,
        *,
        input_files: list[Path],
        output_path: Path | None,
    ) -> SandboxScriptPaths:
        workspace_root = str(
            getattr(booter, "workspace_root", cls._SANDBOX_WORKSPACE_ROOT)
            or cls._SANDBOX_WORKSPACE_ROOT
        )
        exec_relative_dir = cls._SANDBOX_EXEC_ROOT / uuid4().hex
        runtime_input_files: list[str] = []
        remote_input_files: list[str] = []
        for index, original_path in enumerate(input_files):
            relative_input_path = (
                exec_relative_dir / "_inputs" / str(index) / original_path.name
            )
            runtime_input_files.append(
                cls._join_sandbox_path(workspace_root, relative_input_path)
            )
            remote_input_files.append(relative_input_path.as_posix())

        relative_result_path = exec_relative_dir / "result.json"
        runtime_result_path = cls._join_sandbox_path(workspace_root, relative_result_path)

        runtime_output_path: str | None = None
        remote_output_path: str | None = None
        if output_path is not None:
            relative_output_path = exec_relative_dir / "_output" / output_path.name
            runtime_output_path = cls._join_sandbox_path(
                workspace_root,
                relative_output_path,
            )
            remote_output_path = relative_output_path.as_posix()

        return SandboxScriptPaths(
            exec_dir=cls._join_sandbox_path(workspace_root, exec_relative_dir),
            result_path=runtime_result_path,
            remote_result_path=relative_result_path.as_posix(),
            input_files=runtime_input_files,
            remote_input_files=remote_input_files,
            output_path=runtime_output_path,
            remote_output_path=remote_output_path,
        )

    @staticmethod
    def _build_prepare_script(directory_paths: list[str]) -> str:
        unique_paths = list(dict.fromkeys(directory_paths))
        serialized_paths = json.dumps(unique_paths, ensure_ascii=False)
        return textwrap.dedent(
            f"""
            import json
            from pathlib import Path

            for raw_path in json.loads({serialized_paths!r}):
                Path(raw_path).mkdir(parents=True, exist_ok=True)
            """
        ).strip()

    @staticmethod
    def _build_cleanup_script(directory_path: str) -> str:
        serialized_path = json.dumps(directory_path, ensure_ascii=False)
        return textwrap.dedent(
            f"""
            import json
            import shutil
            from pathlib import Path

            shutil.rmtree(Path(json.loads({serialized_path!r})), ignore_errors=True)
            """
        ).strip()

    @staticmethod
    def _format_exec_traceback(exec_result: dict) -> str:
        parts: list[str] = []
        error_text = str(exec_result.get("error", "") or "").strip()
        if error_text:
            parts.append(error_text)
        output_text = str(exec_result.get("output", "") or "").strip()
        if not output_text:
            data = exec_result.get("data")
            if isinstance(data, dict):
                output_payload = data.get("output")
                if isinstance(output_payload, dict):
                    output_text = str(output_payload.get("text", "") or "").strip()
        if output_text and output_text not in parts:
            parts.append(output_text)
        return "\n".join(parts).strip()

    async def _cleanup_sandbox_exec_dir(self, booter, exec_dir: str) -> None:
        try:
            await booter.python.exec(
                self._build_cleanup_script(exec_dir),
                timeout=10,
                silent=True,
            )
        except Exception as exc:  # pragma: no cover - cleanup failure should not mask result
            logger.warning(f"[文件管理] 清理 Excel sandbox 临时目录失败: {exc}")

    async def _run_script_process(
        self,
        *,
        script: str,
        input_files: list[Path],
        output_path: Path | None,
        runtime_mode: str,
        booter=None,
    ) -> ScriptProcessResult:
        if runtime_mode == "sandbox":
            if booter is None:
                return ScriptProcessResult(
                    success=False,
                    mode="error",
                    error="Excel sandbox 不可用",
                    traceback="Missing sandbox booter",
                    script=script,
                    retryable=False,
                )
            return await self._run_sandbox_script_process(
                booter=booter,
                script=script,
                input_files=input_files,
                output_path=output_path,
            )
        return await asyncio.to_thread(
            self._run_local_script_process,
            script=script,
            input_files=input_files,
            output_path=output_path,
        )

    async def _run_sandbox_script_process(
        self,
        *,
        booter,
        script: str,
        input_files: list[Path],
        output_path: Path | None,
    ) -> ScriptProcessResult:
        workspace_root = self._workspace_service.plugin_data_path.resolve()
        with tempfile.TemporaryDirectory(
            prefix="excel_script_",
            dir=str(workspace_root),
        ) as temp_dir:
            temp_root = Path(temp_dir)
            sandbox_paths = self._build_sandbox_paths(
                booter,
                input_files=input_files,
                output_path=output_path,
            )
            result_path = temp_root / "result.json"
            try:
                try:
                    prepare_result = await booter.python.exec(
                        self._build_prepare_script(
                            [
                                str(Path(path).parent)
                                for path in [
                                    *sandbox_paths.input_files,
                                    sandbox_paths.result_path,
                                    sandbox_paths.output_path,
                                ]
                                if path
                            ]
                        ),
                        timeout=10,
                        silent=True,
                    )
                except Exception as exc:
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error="Excel sandbox 初始化失败",
                        traceback=str(exc),
                        script=script,
                    )
                if not prepare_result.get("success", False):
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error="Excel sandbox 初始化失败",
                        traceback=self._format_exec_traceback(prepare_result),
                        script=script,
                    )

                for local_path, remote_path in zip(
                    input_files,
                    sandbox_paths.remote_input_files,
                    strict=True,
                ):
                    try:
                        upload_result = await booter.upload_file(
                            str(local_path),
                            remote_path,
                        )
                    except Exception as exc:
                        return ScriptProcessResult(
                            success=False,
                            mode="error",
                            error="Excel 输入文件上传失败",
                            traceback=str(exc),
                            script=script,
                        )
                    if not upload_result.get("success", False):
                        return ScriptProcessResult(
                            success=False,
                            mode="error",
                            error="Excel 输入文件上传失败",
                            traceback=str(upload_result.get("message", "") or ""),
                            script=script,
                            retryable=False,
                        )

                try:
                    exec_result = await booter.python.exec(
                        self._build_runner_script(
                            script=script,
                            input_files=sandbox_paths.input_files,
                            output_path=sandbox_paths.output_path,
                            result_path=sandbox_paths.result_path,
                        ),
                        timeout=self._SCRIPT_TIMEOUT_SECONDS,
                        silent=True,
                    )
                except Exception as exc:
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error="Excel 脚本执行失败",
                        traceback=str(exc),
                        script=script,
                    )
                if not exec_result.get("success", False):
                    traceback_text = self._format_exec_traceback(exec_result)
                    error_text = "Excel 脚本执行失败"
                    if "timeout" in traceback_text.lower():
                        error_text = "Excel 脚本执行超时"
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error=error_text,
                        traceback=traceback_text,
                        script=script,
                    )

                try:
                    await booter.download_file(
                        sandbox_paths.remote_result_path,
                        str(result_path),
                    )
                except Exception as exc:
                    traceback_text = str(exc)
                    exec_traceback = self._format_exec_traceback(exec_result)
                    if exec_traceback:
                        traceback_text = f"{traceback_text}\n{exec_traceback}".strip()
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error="Excel 脚本未返回结果",
                        traceback=traceback_text,
                        script=script,
                    )

                try:
                    payload = json.loads(result_path.read_text(encoding="utf-8"))
                except (json.JSONDecodeError, OSError) as exc:
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error="Excel 脚本结果解析失败",
                        traceback=str(exc),
                        script=script,
                    )
                if not isinstance(payload, dict):
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error="Excel 脚本结果解析失败",
                        traceback=f"Unexpected JSON payload type: {type(payload).__name__}",
                        script=script,
                    )
                if not payload.get("success"):
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error=str(payload.get("error", "脚本执行失败")),
                        traceback=str(payload.get("traceback", "")),
                        script=script,
                    )

                mode = str(payload.get("mode", ""))
                if mode == "text":
                    return ScriptProcessResult(
                        success=True,
                        mode="text",
                        result_text=str(payload.get("result_text", "")),
                    )
                if mode == "file":
                    output_path_value = payload.get("output_path")
                    if not isinstance(output_path_value, str) or not output_path_value.strip():
                        return ScriptProcessResult(
                            success=False,
                            mode="error",
                            error="脚本返回了 file 模式，但缺少 output_path",
                            traceback="",
                            script=script,
                        )
                    if output_path is None or sandbox_paths.output_path is None:
                        return ScriptProcessResult(
                            success=False,
                            mode="error",
                            error="脚本返回了 file 模式，但当前请求未提供 output_path",
                            traceback="",
                            script=script,
                        )
                    if output_path_value != sandbox_paths.output_path:
                        return ScriptProcessResult(
                            success=False,
                            mode="error",
                            error="脚本返回了无效的 output_path",
                            traceback=f"Unexpected output_path: {output_path_value}",
                            script=script,
                        )
                    try:
                        await booter.download_file(
                            sandbox_paths.remote_output_path or "",
                            str(output_path),
                        )
                    except FileNotFoundError:
                        return ScriptProcessResult(
                            success=False,
                            mode="error",
                            error="脚本返回了 file 模式，但 output_path 不存在",
                            traceback="",
                            script=script,
                        )
                    except Exception as exc:
                        return ScriptProcessResult(
                            success=False,
                            mode="error",
                            error="Excel 脚本结果下载失败",
                            traceback=str(exc),
                            script=script,
                        )
                    if not output_path.exists():
                        return ScriptProcessResult(
                            success=False,
                            mode="error",
                            error="脚本返回了 file 模式，但 output_path 不存在",
                            traceback="",
                            script=script,
                        )
                    return ScriptProcessResult(
                        success=True,
                        mode="file",
                        output_path=output_path,
                    )
                return ScriptProcessResult(
                    success=False,
                    mode="error",
                    error=f"未知脚本返回模式: {mode}",
                    traceback="",
                    script=script,
                )
            finally:
                await self._cleanup_sandbox_exec_dir(booter, sandbox_paths.exec_dir)

    def _run_local_script_process(
        self,
        *,
        script: str,
        input_files: list[Path],
        output_path: Path | None,
    ) -> ScriptProcessResult:
        workspace_root = self._workspace_service.plugin_data_path.resolve()
        with tempfile.TemporaryDirectory(
            prefix="excel_script_",
            dir=str(workspace_root),
        ) as temp_dir:
            temp_root = Path(temp_dir)
            copied_input_files: list[Path] = []
            inputs_root = temp_root / "_inputs"
            inputs_root.mkdir()
            for index, original_path in enumerate(input_files):
                copied_parent = inputs_root / str(index)
                copied_parent.mkdir()
                copied_path = copied_parent / original_path.name
                shutil.copy2(original_path, copied_path)
                copied_input_files.append(copied_path)

            runner_path = temp_root / "runner.py"
            result_path = temp_root / "result.json"
            runner_path.write_text(
                self._build_runner_script(
                    script=script,
                    input_files=[str(path.resolve()) for path in copied_input_files],
                    output_path=(
                        str(output_path.resolve()) if output_path is not None else None
                    ),
                    result_path=str(result_path.resolve()),
                ),
                encoding="utf-8",
            )
            try:
                completed = subprocess.run(
                    [sys.executable, str(runner_path)],
                    cwd=str(workspace_root),
                    capture_output=True,
                    text=True,
                    timeout=self._SCRIPT_TIMEOUT_SECONDS,
                )
            except subprocess.TimeoutExpired as exc:
                return ScriptProcessResult(
                    success=False,
                    mode="error",
                    error="Excel 脚本执行超时",
                    traceback=str(exc),
                    script=script,
                )
            except OSError as exc:
                return ScriptProcessResult(
                    success=False,
                    mode="error",
                    error="Excel 脚本执行失败",
                    traceback=str(exc),
                    script=script,
                )

            if not result_path.exists():
                return ScriptProcessResult(
                    success=False,
                    mode="error",
                    error="Excel 脚本未返回结果",
                    traceback=(completed.stderr or completed.stdout or "").strip(),
                    script=script,
                )

            try:
                payload = json.loads(result_path.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError) as exc:
                return ScriptProcessResult(
                    success=False,
                    mode="error",
                    error="Excel 脚本结果解析失败",
                    traceback=str(exc),
                    script=script,
                )
            if not isinstance(payload, dict):
                return ScriptProcessResult(
                    success=False,
                    mode="error",
                    error="Excel 脚本结果解析失败",
                    traceback=f"Unexpected JSON payload type: {type(payload).__name__}",
                    script=script,
                )
            if not payload.get("success"):
                return ScriptProcessResult(
                    success=False,
                    mode="error",
                    error=str(payload.get("error", "脚本执行失败")),
                    traceback=str(payload.get("traceback", "")),
                    script=script,
                )

            mode = str(payload.get("mode", ""))
            if mode == "text":
                return ScriptProcessResult(
                    success=True,
                    mode="text",
                    result_text=str(payload.get("result_text", "")),
                )
            if mode == "file":
                output_path_value = payload.get("output_path")
                if not isinstance(output_path_value, str) or not output_path_value.strip():
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error="脚本返回了 file 模式，但缺少 output_path",
                        traceback="",
                        script=script,
                    )
                if output_path is None:
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error="脚本返回了 file 模式，但当前请求未提供 output_path",
                        traceback="",
                        script=script,
                    )
                expected_output_path = str(output_path.resolve())
                if output_path_value != expected_output_path:
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error="脚本返回了无效的 output_path",
                        traceback=f"Unexpected output_path: {output_path_value}",
                        script=script,
                    )
                if not output_path.exists():
                    return ScriptProcessResult(
                        success=False,
                        mode="error",
                        error="脚本返回了 file 模式，但 output_path 不存在",
                        traceback="",
                        script=script,
                    )
                return ScriptProcessResult(
                    success=True,
                    mode="file",
                    output_path=output_path,
                )
            return ScriptProcessResult(
                success=False,
                mode="error",
                error=f"未知脚本返回模式: {mode}",
                traceback="",
                script=script,
            )

    @staticmethod
    def _build_runner_script(
        *,
        script: str,
        input_files: list[str],
        output_path: str | None,
        result_path: str,
    ) -> str:
        serialized_input_files = json.dumps(input_files, ensure_ascii=False)
        serialized_output_path = json.dumps(output_path, ensure_ascii=False)
        serialized_result_path = json.dumps(result_path, ensure_ascii=False)
        serialized_script = json.dumps(script, ensure_ascii=False)
        return textwrap.dedent(
            f"""
            import json
            import traceback
            from pathlib import Path

            input_files = [Path(path) for path in json.loads({serialized_input_files!r})]
            output_path_value = json.loads({serialized_output_path!r})
            output_path = Path(output_path_value) if output_path_value else None
            result_path = Path(json.loads({serialized_result_path!r}))
            script = json.loads({serialized_script!r})
            initial_exists = output_path.exists() if output_path is not None else False
            initial_size = output_path.stat().st_size if initial_exists else None
            initial_mtime_ns = output_path.stat().st_mtime_ns if initial_exists else None

            try:
                import openpyxl
                from openpyxl import Workbook, load_workbook

                namespace = {{
                    "__name__": "__main__",
                    "openpyxl": openpyxl,
                    "Workbook": Workbook,
                    "load_workbook": load_workbook,
                    "Path": Path,
                    "input_files": input_files,
                    "output_path": output_path,
                    "result_text": None,
                }}

                exec(script, namespace, namespace)
                result_text = namespace.get("result_text")
                has_text = result_text is not None
                has_file = False
                if output_path is not None and output_path.exists():
                    if not initial_exists:
                        has_file = True
                    else:
                        has_file = (
                            output_path.stat().st_size != initial_size
                            or output_path.stat().st_mtime_ns != initial_mtime_ns
                        )
                if has_text and has_file:
                    payload = {{
                        "success": False,
                        "error": "脚本不能同时设置 result_text 并写出 output_path",
                        "traceback": "",
                    }}
                elif not has_text and not has_file:
                    payload = {{
                        "success": False,
                        "error": "脚本执行完成，但既没有设置 result_text，也没有写出 output_path",
                        "traceback": "",
                    }}
                elif has_text:
                    payload = {{
                        "success": True,
                        "mode": "text",
                        "result_text": str(result_text),
                    }}
                else:
                    payload = {{
                        "success": True,
                        "mode": "file",
                        "output_path": str(output_path),
                    }}
            except Exception as exc:
                payload = {{
                    "success": False,
                    "error": str(exc),
                    "traceback": traceback.format_exc(),
                }}

            result_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
            """
        ).strip()


__all__ = ["ExcelScriptService", "ScriptProcessResult"]
