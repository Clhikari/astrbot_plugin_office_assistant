from __future__ import annotations

import json
import textwrap


def build_prepare_script(file_paths: list[str]) -> str:
    unique_paths = list(dict.fromkeys(file_paths))
    serialized_paths = json.dumps(unique_paths, ensure_ascii=False)
    return textwrap.dedent(
        f"""
        import json
        from pathlib import Path

        for raw_path in json.loads({serialized_paths!r}):
            Path(raw_path).parent.mkdir(parents=True, exist_ok=True)
        """
    ).strip()


def build_cleanup_script(directory_path: str) -> str:
    serialized_path = json.dumps(directory_path, ensure_ascii=False)
    return textwrap.dedent(
        f"""
        import json
        import shutil
        from pathlib import Path

        shutil.rmtree(Path(json.loads({serialized_path!r})), ignore_errors=True)
        """
    ).strip()


def build_runner_script(
    *,
    script: str,
    exec_dir: str,
    input_files: list[str],
    output_path: str | None,
    result_path: str,
    helper_script: str | None = None,
) -> str:
    serialized_exec_dir = json.dumps(exec_dir, ensure_ascii=False)
    serialized_input_files = json.dumps(input_files, ensure_ascii=False)
    serialized_output_path = json.dumps(output_path, ensure_ascii=False)
    serialized_result_path = json.dumps(result_path, ensure_ascii=False)
    serialized_script = json.dumps(script, ensure_ascii=False)
    serialized_helper_script = json.dumps(
        helper_script or build_script_helper_template(),
        ensure_ascii=False,
    )
    return textwrap.dedent(
        f"""
        import json
        import os
        import shutil
        import traceback
        from pathlib import Path

        workspace_root = Path.cwd()
        exec_dir = Path(json.loads({serialized_exec_dir!r}))
        if not exec_dir.is_absolute():
            exec_dir = workspace_root / exec_dir
        exec_dir = exec_dir.resolve()
        script_input_files = [
            Path(path) for path in json.loads({serialized_input_files!r})
        ]
        input_files = [
            path if path.is_absolute() else (exec_dir / path).resolve()
            for path in script_input_files
        ]
        output_path_value = json.loads({serialized_output_path!r})
        reported_output_path = output_path_value
        output_path = Path(output_path_value) if output_path_value else None
        if output_path is not None and not output_path.is_absolute():
            output_path = (workspace_root / output_path).resolve()
        result_path = Path(json.loads({serialized_result_path!r}))
        if not result_path.is_absolute():
            result_path = (workspace_root / result_path).resolve()
        script = json.loads({serialized_script!r})
        helper_script = json.loads({serialized_helper_script!r})
        initial_exists = output_path.exists() if output_path is not None else False
        initial_size = output_path.stat().st_size if initial_exists else None
        initial_mtime_ns = output_path.stat().st_mtime_ns if initial_exists else None

        try:
            import openpyxl
            from openpyxl import Workbook, load_workbook

            namespace = {{
                "__name__": "__main__",
                "openpyxl": openpyxl,
                "Workbook": Workbook,
                "load_workbook": load_workbook,
                "Path": Path,
                "input_files": script_input_files,
                "output_path": output_path,
                "result_text": None,
            }}

            original_cwd = Path.cwd()
            try:
                os.chdir(exec_dir)
                exec(helper_script, namespace, namespace)
                exec(script, namespace, namespace)
            finally:
                os.chdir(original_cwd)

            if output_path is not None and not output_path.exists():
                output_alias_path = exec_dir / output_path.name
                if output_alias_path.exists():
                    shutil.copy2(output_alias_path, output_path)
            if output_path is not None and not output_path.exists():
                known_input_names = {{path.name for path in input_files}}
                candidates = []
                for candidate in exec_dir.iterdir():
                    if not candidate.is_file():
                        continue
                    if candidate.resolve() == result_path.resolve():
                        continue
                    if candidate.resolve() == output_path.resolve():
                        continue
                    if candidate.name in known_input_names:
                        continue
                    if candidate.suffix.lower() in {{".xlsx", ".xlsm", ".xls"}}:
                        candidates.append(candidate)
                if len(candidates) == 1:
                    shutil.copy2(candidates[0], output_path)
            output_changed = False
            if output_path is not None and output_path.exists():
                if not initial_exists:
                    output_changed = True
                else:
                    output_changed = (
                        output_path.stat().st_size != initial_size
                        or output_path.stat().st_mtime_ns != initial_mtime_ns
                    )
            postprocess_output = namespace.get(
                "_office_assistant_postprocess_output_file"
            )
            if output_path is not None and output_changed and callable(postprocess_output):
                current_cwd = Path.cwd()
                try:
                    os.chdir(exec_dir)
                    postprocess_output(output_path)
                finally:
                    os.chdir(current_cwd)

            result_text = namespace.get("result_text")
            has_text = result_text is not None
            has_file = False
            if output_path is not None and output_path.exists():
                if not initial_exists:
                    has_file = True
                else:
                    has_file = (
                        output_path.stat().st_size != initial_size
                        or output_path.stat().st_mtime_ns != initial_mtime_ns
                    )
            if has_text and has_file:
                payload = {{
                    "success": False,
                    "error": "脚本不能同时设置 result_text 并写出 output_path",
                    "traceback": "",
                }}
            elif not has_text and not has_file:
                payload = {{
                    "success": False,
                    "error": "脚本执行完成，但既没有设置 result_text，也没有写出 output_path",
                    "traceback": "",
                }}
            elif has_text:
                payload = {{
                    "success": True,
                    "mode": "text",
                    "result_text": str(result_text),
                }}
            else:
                payload = {{
                    "success": True,
                    "mode": "file",
                    "output_path": reported_output_path,
                }}
        except SyntaxError as exc:
            line_text = (exc.text or "").strip()
            location = f"第 {{exc.lineno}} 行" if exc.lineno else "未知行"
            payload = {{
                "success": False,
                "error": (
                    f"脚本语法错误（{{location}}）：{{exc.msg}}。"
                    f"出错代码：{{line_text}}"
                ),
                "traceback": traceback.format_exc(),
            }}
        except Exception as exc:
            payload = {{
                "success": False,
                "error": str(exc),
                "traceback": traceback.format_exc(),
            }}

        result_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        """
    ).strip()


def build_script_helper_template() -> str:
    return textwrap.dedent(
        r"""
        from copy import copy as _excel_copy
        from openpyxl.styles import Alignment, Border, Font, PatternFill
        from openpyxl.utils import get_column_letter


        def load_input_workbook(index=0, **kwargs):
            return load_workbook(input_files[index], **kwargs)


        def save_output_workbook(workbook, *, auto_format=True):
            if auto_format:
                auto_format_workbook(workbook)
            _request_formula_recalculation(workbook)
            _office_assistant_original_workbook_save(workbook, output_path)


        def _excel_alignment_with(
            alignment,
            *,
            wrap_text=None,
            horizontal=None,
            vertical=None,
        ):
            return Alignment(
                horizontal=horizontal
                if horizontal is not None
                else alignment.horizontal,
                vertical=vertical if vertical is not None else alignment.vertical,
                text_rotation=alignment.text_rotation,
                wrap_text=wrap_text
                if wrap_text is not None
                else alignment.wrap_text,
                shrink_to_fit=alignment.shrink_to_fit,
                indent=alignment.indent,
                relativeIndent=alignment.relativeIndent,
                justifyLastLine=alignment.justifyLastLine,
                readingOrder=alignment.readingOrder,
            )


        def ensure_readable_sheet(
            worksheet,
            *,
            min_column_width=13,
            max_column_width=40,
            min_row_height=18,
        ):
            default_row_height = 15
            row_height_targets = {}

            def display_bounds(cell):
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate not in merged_range:
                        continue
                    if (
                        cell.row != merged_range.min_row
                        or cell.column != merged_range.min_col
                    ):
                        return None
                    return (
                        merged_range.min_row,
                        merged_range.min_col,
                        merged_range.max_row,
                        merged_range.max_col,
                    )
                return (cell.row, cell.column, cell.row, cell.column)

            def current_row_height(row_index):
                return worksheet.row_dimensions[row_index].height or default_row_height

            def _display_line_width(value):
                return sum(2 if ord(char) > 127 else 1 for char in str(value))

            def current_column_width(column_index):
                column_letter = get_column_letter(column_index)
                return (
                    worksheet.column_dimensions[column_letter].width
                    or min_column_width
                )

            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str) or not value.strip():
                        continue
                    bounds = display_bounds(cell)
                    if bounds is None:
                        continue
                    lines = [
                        line
                        for line in value.replace("\r\n", "\n")
                        .replace("\r", "\n")
                        .split("\n")
                        if line
                    ]
                    if not lines:
                        continue

                    max_line_length = max(_display_line_width(line) for line in lines)
                    if len(lines) > 1 or max_line_length > 20:
                        cell.alignment = _excel_alignment_with(
                            cell.alignment,
                            wrap_text=True,
                            vertical=cell.alignment.vertical or "center",
                        )
                    else:
                        continue

                    min_row, min_col, max_row, max_col = bounds
                    span_rows = max(max_row - min_row + 1, 1)
                    available_width = sum(
                        current_column_width(column_index)
                        for column_index in range(min_col, max_col + 1)
                    )
                    estimated_line_count = max(
                        len(lines),
                        *[
                            max(
                                1,
                                int(
                                    (
                                        _display_line_width(line)
                                        + max(int(available_width), 1)
                                        - 1
                                    )
                                    / max(int(available_width), 1)
                                ),
                            )
                            for line in lines
                        ],
                    )
                    target_total_height = min(
                        max(min_row_height * span_rows, 15 * estimated_line_count),
                        120,
                    )
                    current_total_height = sum(
                        current_row_height(row_index)
                        for row_index in range(min_row, max_row + 1)
                    )
                    uniform_height = max(
                        min_row_height,
                        max(target_total_height, current_total_height) / span_rows,
                    )
                    for row_index in range(min_row, max_row + 1):
                        row_height_targets[row_index] = max(
                            row_height_targets.get(row_index, 0),
                            uniform_height,
                        )

            for row_index, height in row_height_targets.items():
                worksheet.row_dimensions[row_index].height = height

            for column_index in range(1, worksheet.max_column + 1):
                column_letter = get_column_letter(column_index)
                values = [
                    worksheet.cell(row=row_index, column=column_index).value
                    for row_index in range(1, worksheet.max_row + 1)
                ]
                max_length = max(
                    (
                        max(
                            _display_line_width(line)
                            for line in str(value).splitlines() or [""]
                        )
                        for value in values
                        if value not in (None, "")
                    ),
                    default=0,
                )
                if max_length == 0:
                    continue
                target_width = min(
                    max(max_length + 2, min_column_width),
                    max_column_width,
                )
                dimension = worksheet.column_dimensions[column_letter]
                current_width = dimension.width or 0
                if current_width < target_width:
                    dimension.width = target_width
            return worksheet


        def format_table_sheet(
            worksheet,
            *,
            widths=None,
            header_row=1,
            freeze_panes=True,
            auto_filter=True,
        ):
            header_fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAF7",
            )
            header_font = Font(bold=True)
            for cell in worksheet[header_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = _excel_alignment_with(
                    cell.alignment,
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

            for row in worksheet.iter_rows(min_row=header_row + 1):
                for cell in row:
                    cell.alignment = _excel_alignment_with(
                        cell.alignment,
                        vertical=cell.alignment.vertical or "center",
                        wrap_text=True,
                    )

            if freeze_panes is True:
                worksheet.freeze_panes = f"A{header_row + 1}"
            elif freeze_panes:
                worksheet.freeze_panes = freeze_panes
            if auto_filter and worksheet.max_row >= 1 and worksheet.max_column >= 1:
                last_column = get_column_letter(worksheet.max_column)
                worksheet.auto_filter.ref = f"A{header_row}:{last_column}{worksheet.max_row}"
            ensure_readable_sheet(worksheet)

            for key, width in (widths or {}).items():
                if isinstance(key, int):
                    column_letter = get_column_letter(key)
                else:
                    column_letter = str(key)
                worksheet.column_dimensions[column_letter].width = width
            return worksheet


        def format_course_list_sheet(worksheet):
            return format_table_sheet(
                worksheet,
                widths={
                    "A": 16,
                    "B": 12,
                    "C": 14,
                    "D": 14,
                    "E": 28,
                    "F": 16,
                    "G": 20,
                    "H": 18,
                },
            )


        def _input_workbook_sheet_names():
            sheet_names = set()
            for index, _path in enumerate(input_files):
                try:
                    workbook = load_input_workbook(index, read_only=True)
                except Exception:
                    continue
                try:
                    sheet_names.update(workbook.sheetnames)
                finally:
                    workbook.close()
            return sheet_names


        def _detect_table_header_row(worksheet):
            if worksheet.max_row < 2 or worksheet.max_column < 2:
                return None
            known_headers = {
                "class",
                "classname",
                "course",
                "department",
                "examdate",
                "examtime",
                "issue",
                "note",
                "rawnote",
                "room",
                "seatno",
                "teacher",
                "toconfirm",
                "totalexams",
                "weekday",
                "weeks",
            }
            first_candidate = None
            for row_index in range(1, min(worksheet.max_row, 8) + 1):
                values = [
                    worksheet.cell(row=row_index, column=column_index).value
                    for column_index in range(1, worksheet.max_column + 1)
                ]
                normalized_values = [
                    str(value).strip().lower().replace(" ", "")
                    for value in values
                    if value not in (None, "")
                ]
                if len(normalized_values) < 2:
                    continue
                if first_candidate is None:
                    first_candidate = row_index
                if any(value in known_headers for value in normalized_values):
                    return row_index
            return first_candidate


        def _request_formula_recalculation(workbook):
            calculation = getattr(workbook, "calculation", None)
            if calculation is None:
                return workbook
            calculation.fullCalcOnLoad = True
            calculation.forceFullCalc = True
            calculation.calcMode = "auto"
            return workbook


        def auto_format_workbook(workbook, *, skip_sheet_names=None):
            input_sheet_names = (
                set(skip_sheet_names)
                if skip_sheet_names is not None
                else _input_workbook_sheet_names()
            )
            for worksheet in workbook.worksheets:
                header_row = _detect_table_header_row(worksheet)
                if worksheet.title in input_sheet_names:
                    ensure_readable_sheet(worksheet)
                elif header_row is not None:
                    format_table_sheet(worksheet, header_row=header_row)
                else:
                    ensure_readable_sheet(worksheet)
            return workbook


        if not hasattr(Workbook, "_office_assistant_original_save"):
            Workbook._office_assistant_original_save = Workbook.save
        _office_assistant_original_workbook_save = (
            Workbook._office_assistant_original_save
        )


        def _office_assistant_matches_output_path(filename):
            if output_path is None:
                return False
            try:
                filename_path = Path(filename)
            except TypeError:
                return False
            try:
                return filename_path.resolve() == output_path.resolve()
            except (OSError, RuntimeError, ValueError):
                return str(filename_path) == str(output_path)


        def _office_assistant_postprocess_workbook(workbook):
            auto_format_workbook(workbook)
            _request_formula_recalculation(workbook)
            return workbook


        def _office_assistant_postprocess_output_file(path=None):
            target_path = Path(path or output_path) if path or output_path else None
            if target_path is None:
                return None
            if target_path.suffix.lower() not in {".xlsx", ".xlsm"}:
                return None
            try:
                workbook = load_workbook(target_path)
            except Exception:
                return None
            try:
                _office_assistant_postprocess_workbook(workbook)
                _office_assistant_original_workbook_save(workbook, target_path)
                return workbook
            finally:
                workbook.close()


        def _office_assistant_save_workbook(workbook, filename):
            if (
                _office_assistant_matches_output_path(filename)
                and not getattr(workbook, "_office_assistant_saving_output", False)
            ):
                workbook._office_assistant_saving_output = True
                try:
                    _office_assistant_postprocess_workbook(workbook)
                    return _office_assistant_original_workbook_save(
                        workbook,
                        filename,
                    )
                finally:
                    try:
                        del workbook._office_assistant_saving_output
                    except AttributeError:
                        pass
            return _office_assistant_original_workbook_save(workbook, filename)


        Workbook.save = _office_assistant_save_workbook


        def _copy_cell_style(source_cell, target_cell):
            if source_cell.has_style:
                target_cell._style = _excel_copy(source_cell._style)
            if source_cell.hyperlink:
                target_cell._hyperlink = _excel_copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = _excel_copy(source_cell.comment)


        def copy_sheet_layout(source_sheet, target_sheet, *, copy_cell_styles=True):
            for key, dimension in source_sheet.column_dimensions.items():
                target_dimension = target_sheet.column_dimensions[key]
                target_dimension.width = dimension.width
                target_dimension.hidden = dimension.hidden
                target_dimension.outlineLevel = dimension.outlineLevel
                target_dimension.collapsed = dimension.collapsed

            for key, dimension in source_sheet.row_dimensions.items():
                target_dimension = target_sheet.row_dimensions[key]
                target_dimension.height = dimension.height
                target_dimension.hidden = dimension.hidden
                target_dimension.outlineLevel = dimension.outlineLevel
                target_dimension.collapsed = dimension.collapsed

            target_sheet.freeze_panes = source_sheet.freeze_panes
            target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
            target_sheet.sheet_properties = _excel_copy(source_sheet.sheet_properties)
            target_sheet.page_margins = _excel_copy(source_sheet.page_margins)
            target_sheet.page_setup = _excel_copy(source_sheet.page_setup)

            existing_ranges = {
                str(merged_range) for merged_range in target_sheet.merged_cells.ranges
            }
            for merged_range in source_sheet.merged_cells.ranges:
                range_text = str(merged_range)
                if range_text in existing_ranges:
                    continue
                try:
                    target_sheet.merge_cells(range_text)
                except ValueError:
                    pass

            if copy_cell_styles:
                for row in source_sheet.iter_rows():
                    for source_cell in row:
                        _copy_cell_style(
                            source_cell,
                            target_sheet[source_cell.coordinate],
                        )
            return target_sheet


        def preserve_input_sheet_layout(
            workbook,
            *,
            input_index=0,
            sheet_names=None,
            copy_cell_styles=True,
        ):
            source_workbook = load_input_workbook(input_index)
            try:
                names = sheet_names or source_workbook.sheetnames
                for sheet_name in names:
                    if (
                        sheet_name in source_workbook.sheetnames
                        and sheet_name in workbook.sheetnames
                    ):
                        copy_sheet_layout(
                            source_workbook[sheet_name],
                            workbook[sheet_name],
                            copy_cell_styles=copy_cell_styles,
                        )
                return workbook
            finally:
                source_workbook.close()
        """
    ).strip()


__all__ = [
    "build_cleanup_script",
    "build_prepare_script",
    "build_runner_script",
    "build_script_helper_template",
]
