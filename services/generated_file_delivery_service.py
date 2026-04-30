from collections.abc import Sequence
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal

from astrbot.api import logger
from astrbot.api.event import AstrMessageEvent


@dataclass(slots=True)
class GeneratedFileDeliveryResult:
    status: Literal["sent", "missing", "oversized", "invalid", "review_required"]
    file_size: int = 0
    max_size: int = 0
    validation_errors: list[str] | None = None
    quality_summary: dict[str, Any] | None = None


class GeneratedFileDeliveryService:
    _EXCEL_FORMULA_SUFFIXES = {".xlsx", ".xlsm"}
    _FORMULA_DIVISOR_RE = re.compile(
        r"/\s*((?:(?:'(?:[^']|'')+'|[A-Za-z_][A-Za-z0-9_.]*)!)?\$?[A-Z]{1,3}\$?\d+)"
    )
    _DOUBLE_QUOTED_TEXT_RE = re.compile(r'"(?:[^"]|"")*"')
    _SINGLE_QUOTED_LITERAL_RE = re.compile(r"(?<![A-Za-z0-9_])'([^']+)'(?!\s*!)")
    _MAX_VALIDATION_ERRORS = 8
    _CLEAN_TEXT_SHEET_MARKERS = ("clean", "import", "清洗", "导入")
    _PIE_CHART_CLASS_NAMES = {"DoughnutChart", "PieChart", "ProjectedPieChart"}

    def __init__(self, *, workspace_service, delivery_service) -> None:
        self._workspace_service = workspace_service
        self._delivery_service = delivery_service

    @classmethod
    def _normalize_cell_reference(cls, cell_reference: str) -> str:
        return cell_reference.replace("$", "").upper()

    @classmethod
    def _normalize_formula_reference(cls, reference: str) -> str:
        return reference.replace("$", "").replace("'", "").replace(" ", "").upper()

    @classmethod
    def _split_formula_reference(cls, reference: str) -> tuple[str | None, str]:
        normalized = reference.replace("$", "")
        if "!" not in normalized:
            return None, cls._normalize_cell_reference(normalized)

        sheet_name, cell_reference = normalized.rsplit("!", 1)
        sheet_name = sheet_name.strip()
        if sheet_name.startswith("'") and sheet_name.endswith("'"):
            sheet_name = sheet_name[1:-1].replace("''", "'")
        return sheet_name, cls._normalize_cell_reference(cell_reference)

    @staticmethod
    def _format_formula_reference(
        sheet_name: str | None,
        cell_reference: str,
    ) -> str:
        if sheet_name:
            return f"{sheet_name}!{cell_reference}"
        return cell_reference

    @staticmethod
    def _normalize_header(header: Any) -> str:
        return re.sub(r"[\s_]+", "", str(header or "").strip().lower())

    @classmethod
    def _find_header_column(
        cls,
        headers: list[Any],
        aliases: tuple[str, ...],
    ) -> int | None:
        normalized_aliases = {cls._normalize_header(alias) for alias in aliases}
        for index, header in enumerate(headers, start=1):
            if cls._normalize_header(header) in normalized_aliases:
                return index
        return None

    @classmethod
    def _count_repeated_header_rows(cls, worksheet, headers: list[Any]) -> int:
        normalized_headers = [cls._normalize_header(header) for header in headers]
        repeated_rows = 0
        for row_index in range(2, worksheet.max_row + 1):
            normalized_values = [
                cls._normalize_header(
                    worksheet.cell(row=row_index, column=column).value
                )
                for column in range(1, worksheet.max_column + 1)
            ]
            if not normalized_values or normalized_values[0] != normalized_headers[0]:
                continue
            matched_cells = sum(
                1
                for value, header in zip(normalized_values, normalized_headers)
                if value and value == header
            )
            if matched_cells >= 3:
                repeated_rows += 1
        return repeated_rows

    @classmethod
    def _count_non_formula_values(cls, worksheet, column: int) -> int:
        count = 0
        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=column).value
            if value in (None, ""):
                continue
            if not (isinstance(value, str) and value.startswith("=")):
                count += 1
        return count

    @classmethod
    def _count_cells_with_line_breaks_or_tabs(cls, worksheet) -> int:
        count = 0
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                if "\n" in value or "\r" in value or "\t" in value:
                    count += 1
        return count

    @classmethod
    def _sheet_expects_clean_text(cls, sheet_name: str) -> bool:
        normalized_name = cls._normalize_header(sheet_name)
        return any(
            marker in normalized_name for marker in cls._CLEAN_TEXT_SHEET_MARKERS
        )

    @staticmethod
    def _count_conditional_formatting_rules(worksheet) -> int:
        conditional_formatting = getattr(worksheet, "conditional_formatting", None)
        if conditional_formatting is None:
            return 0

        rule_map = getattr(conditional_formatting, "_cf_rules", None)
        if rule_map is not None:
            try:
                return sum(len(rules) for rules in rule_map.values())
            except TypeError:
                pass

        try:
            return sum(
                len(getattr(item, "rules", ()) or ()) for item in conditional_formatting
            )
        except TypeError:
            pass

        try:
            return len(conditional_formatting)
        except TypeError:
            return 0

    @classmethod
    def _formula_guards_divisor(cls, formula: str, cell_reference: str) -> bool:
        reference = cls._normalize_formula_reference(cell_reference)
        normalized = cls._normalize_formula_reference(formula)
        non_zero_checks = (
            f"{reference}<>0",
            f"0<>{reference}",
            f"{reference}!=0",
            f"0!={reference}",
            f"{reference}>0",
            f"0<{reference}",
            f"{reference}<0",
            f"0>{reference}",
        )
        return (
            f"{reference}=0" in normalized
            or f'{reference}=""' in normalized
            or f"ISBLANK({reference})" in normalized
            or any(check in normalized for check in non_zero_checks)
            or "IFERROR(" in normalized
        )

    @classmethod
    def _find_single_quoted_formula_literal(cls, formula: str) -> str | None:
        formula_without_text = cls._DOUBLE_QUOTED_TEXT_RE.sub('""', formula)
        match = cls._SINGLE_QUOTED_LITERAL_RE.search(formula_without_text)
        if match is None:
            return None
        return match.group(0)

    @classmethod
    def _validate_excel_formula_risks(cls, output_path: Path) -> list[str]:
        if output_path.suffix.lower() not in cls._EXCEL_FORMULA_SUFFIXES:
            return []

        try:
            from openpyxl import load_workbook
        except ImportError:
            return []

        try:
            workbook = load_workbook(output_path, data_only=False, read_only=False)
        except Exception as exc:
            return [f"无法打开生成的 Excel 文件进行校验：{exc}"]

        errors: list[str] = []
        try:
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows():
                    for cell in row:
                        formula = cell.value
                        if not isinstance(formula, str) or not formula.startswith("="):
                            continue
                        single_quoted_literal = cls._find_single_quoted_formula_literal(
                            formula
                        )
                        if single_quoted_literal is not None:
                            errors.append(
                                (
                                    f"{worksheet.title}!{cell.coordinate} 公式 "
                                    f"{formula} 使用了单引号文本 "
                                    f"{single_quoted_literal}，Excel 公式文本必须用双引号"
                                )
                            )
                            if len(errors) >= cls._MAX_VALIDATION_ERRORS:
                                return errors
                        for match in cls._FORMULA_DIVISOR_RE.finditer(formula):
                            divisor_reference = match.group(1)
                            if cls._formula_guards_divisor(
                                formula,
                                divisor_reference,
                            ):
                                continue
                            sheet_name, divisor_ref = cls._split_formula_reference(
                                divisor_reference
                            )
                            divisor_sheet = worksheet
                            if sheet_name:
                                try:
                                    divisor_sheet = workbook[sheet_name]
                                except KeyError:
                                    errors.append(
                                        (
                                            f"{worksheet.title}!{cell.coordinate} 公式 "
                                            f"{formula} 引用了不存在的 Sheet：{sheet_name}"
                                        )
                                    )
                                    if len(errors) >= cls._MAX_VALIDATION_ERRORS:
                                        return errors
                                    continue
                            divisor_value = divisor_sheet[divisor_ref].value
                            if divisor_value in (None, "", 0):
                                display_ref = cls._format_formula_reference(
                                    sheet_name,
                                    divisor_ref,
                                )
                                errors.append(
                                    (
                                        f"{worksheet.title}!{cell.coordinate} 公式 "
                                        f"{formula} 的分母 {display_ref} 为空或 0"
                                    )
                                )
                                if len(errors) >= cls._MAX_VALIDATION_ERRORS:
                                    return errors
        finally:
            workbook.close()
        return errors

    @classmethod
    def _build_sheet_quality_summary(
        cls,
        worksheet,
        *,
        data_row_count: int,
        chart_count: int,
        conditional_formatting_count: int,
    ) -> dict[str, Any]:
        return {
            "name": worksheet.title,
            "rows": worksheet.max_row,
            "data_rows": data_row_count,
            "columns": worksheet.max_column,
            "charts": chart_count,
            "conditional_formatting_rules": conditional_formatting_count,
        }

    @staticmethod
    def _count_sheet_formulas(worksheet) -> int:
        count = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    count += 1
        return count

    @classmethod
    def _quality_warnings_requiring_review(
        cls,
        quality_summary: dict[str, Any] | None,
    ) -> list[str]:
        if quality_summary is None:
            return []
        return list(quality_summary.get("warnings", []))

    @staticmethod
    def _extract_chart_reference_formula(reference) -> str | None:
        if reference is None:
            return None
        for ref_attr in ("numRef", "strRef"):
            ref = getattr(reference, ref_attr, None)
            formula = getattr(ref, "f", None)
            if isinstance(formula, str) and formula.strip():
                return formula.strip()
        return None

    @staticmethod
    def _count_reference_cells(formula: str) -> int | None:
        try:
            from openpyxl.utils.cell import range_boundaries
        except ImportError:
            return None

        range_text = formula.rsplit("!", 1)[-1].strip().replace("$", "")
        if "," in range_text:
            return None
        try:
            min_col, min_row, max_col, max_row = range_boundaries(range_text)
        except ValueError:
            return None
        return (max_col - min_col + 1) * (max_row - min_row + 1)

    @classmethod
    def _collect_chart_warnings(cls, worksheet) -> list[str]:
        warnings: list[str] = []
        for chart in getattr(worksheet, "_charts", []) or []:
            if type(chart).__name__ not in cls._PIE_CHART_CLASS_NAMES:
                continue

            series = list(getattr(chart, "series", []) or [])
            if len(series) != 1:
                warnings.append(
                    f"{worksheet.title} 的饼图包含 {len(series)} 个数据系列，"
                    "可能只会渲染第一个系列"
                )
                continue

            chart_series = series[0]
            category_formula = cls._extract_chart_reference_formula(
                getattr(chart_series, "cat", None)
            )
            value_formula = cls._extract_chart_reference_formula(
                getattr(chart_series, "val", None)
            )
            if category_formula is None or value_formula is None:
                continue

            category_count = cls._count_reference_cells(category_formula)
            value_count = cls._count_reference_cells(value_formula)
            if (
                category_count is not None
                and value_count is not None
                and category_count != value_count
            ):
                warnings.append(
                    f"{worksheet.title} 的饼图分类数量 {category_count} 与"
                    f"数据点数量 {value_count} 不一致"
                )
        return warnings

    @staticmethod
    def _collect_dashboard_warnings(
        worksheet,
        *,
        data_row_count: int,
        chart_count: int,
    ) -> list[str]:
        if worksheet.title.lower() != "dashboard" or data_row_count != 0:
            return []
        if chart_count:
            return ["Dashboard 只有图表，没有数据表或关键指标区"]
        return ["Dashboard 没有数据行或图表"]

    @classmethod
    def _collect_cleaned_data_warnings(cls, worksheet, headers: list[Any]) -> list[str]:
        if worksheet.title.lower() != "cleaneddata":
            return []
        key_column = cls._find_header_column(headers, ("OrderID", "Order ID", "ID"))
        if key_column is None:
            return []

        blank_keys = 0
        for row_index in range(2, worksheet.max_row + 1):
            key_value = worksheet.cell(row=row_index, column=key_column).value
            if key_value in (None, ""):
                blank_keys += 1
        if not blank_keys:
            return []
        return [f"CleanedData 有 {blank_keys} 行主键为空"]

    @classmethod
    def _collect_gross_margin_warnings(
        cls,
        worksheet,
        headers: list[Any],
        *,
        conditional_formatting_count: int,
    ) -> list[str]:
        gross_margin_column = cls._find_header_column(
            headers,
            ("GrossMargin", "Gross Margin"),
        )
        if gross_margin_column is None or conditional_formatting_count != 0:
            return []
        return [f"{worksheet.title} 包含 GrossMargin 列，但没有条件格式规则"]

    @classmethod
    def _collect_issues_sheet_state(
        cls,
        worksheet,
        headers: list[Any],
        *,
        data_row_count: int,
    ) -> tuple[int | None, dict[str, int], list[str]]:
        if worksheet.title.lower() != "issues":
            return None, {}, []

        warnings: list[str] = []
        issue_type_counts: dict[str, int] = {}
        issue_type_column = cls._find_header_column(
            headers,
            ("IssueType", "Issue Type", "Type"),
        )
        if issue_type_column is not None:
            for row_index in range(2, worksheet.max_row + 1):
                issue_type = str(
                    worksheet.cell(row=row_index, column=issue_type_column).value or ""
                ).strip()
                if issue_type:
                    issue_type_counts[issue_type] = (
                        issue_type_counts.get(issue_type, 0) + 1
                    )
        return data_row_count, issue_type_counts, warnings

    @classmethod
    def _collect_product_detail_warnings(
        cls,
        worksheet,
        headers: list[Any],
    ) -> list[str]:
        if worksheet.title.lower() != "productdetail":
            return []

        missing_dimensions = [
            dimension
            for dimension, aliases in (
                ("Region", ("Region",)),
                ("Product", ("Product",)),
                ("Month", ("Month",)),
            )
            if cls._find_header_column(headers, aliases) is None
        ]
        if not missing_dimensions:
            return []
        return ["ProductDetail 缺少维度列：" + "、".join(missing_dimensions)]

    @classmethod
    def _collect_inventory_impact_warnings(
        cls,
        worksheet,
        headers: list[Any],
    ) -> list[str]:
        if worksheet.title.lower() != "inventoryimpact":
            return []

        ending_stock_column = cls._find_header_column(
            headers,
            ("EndingStock", "Ending Stock"),
        )
        if ending_stock_column is None:
            return []
        fixed_values = cls._count_non_formula_values(worksheet, ending_stock_column)
        if not fixed_values:
            return []
        return [f"InventoryImpact 有 {fixed_values} 行 EndingStock 不是公式"]

    @classmethod
    def _collect_summary_sheet_warnings(
        cls,
        worksheet,
        headers: list[Any],
    ) -> list[str]:
        if worksheet.title.lower() != "summary":
            return []

        warnings: list[str] = []
        for column_name, aliases in (
            ("CompletionRate", ("CompletionRate", "Completion Rate")),
            ("Status", ("Status",)),
        ):
            formula_column = cls._find_header_column(headers, aliases)
            if formula_column is None:
                continue
            fixed_values = cls._count_non_formula_values(worksheet, formula_column)
            if fixed_values:
                warnings.append(f"Summary 有 {fixed_values} 行 {column_name} 不是公式")

        target_column = cls._find_header_column(
            headers,
            ("TargetAmount", "Target Amount", "Target"),
        )
        if target_column is None:
            return warnings

        actual_column = cls._find_header_column(
            headers,
            ("ActualAmount", "Actual Amount", "FinalAmount", "Amount"),
        )
        invalid_targets = []
        for row_index in range(2, worksheet.max_row + 1):
            target_value = worksheet.cell(row=row_index, column=target_column).value
            actual_value = (
                worksheet.cell(row=row_index, column=actual_column).value
                if actual_column is not None
                else None
            )
            if target_value in (None, "") or (
                target_value == 0 and actual_value not in (None, "", 0)
            ):
                invalid_targets.append(row_index)
        if invalid_targets:
            warnings.append(f"Summary 有 {len(invalid_targets)} 行目标为空或 0")
        return warnings

    @classmethod
    def _collect_course_list_warnings(
        cls,
        worksheet,
        headers: list[Any],
    ) -> list[str]:
        if worksheet.title.lower() != "courselist":
            return []

        field_aliases = {
            "Class": ("Class", "班级"),
            "Weekday": ("Weekday", "Day", "星期", "周"),
            "SectionStart": ("SectionStart", "Section Start", "StartSection"),
            "SectionEnd": ("SectionEnd", "Section End", "EndSection"),
            "Course": ("Course", "课程"),
            "Teacher": ("Teacher", "老师", "教师"),
            "Room": ("Room", "Classroom", "教室", "地点"),
            "Weeks": ("Weeks", "Week", "周次"),
        }
        columns = {
            field: cls._find_header_column(headers, aliases)
            for field, aliases in field_aliases.items()
        }
        missing_headers = [field for field, column in columns.items() if column is None]
        if missing_headers:
            return ["CourseList 缺少字段：" + "、".join(missing_headers)]

        note_rows: list[int] = []
        incomplete_rows: list[int] = []
        required_fields = tuple(field_aliases)
        for row_index in range(2, worksheet.max_row + 1):
            row_values = [
                worksheet.cell(row=row_index, column=column).value
                for column in range(1, worksheet.max_column + 1)
            ]
            if all(value in (None, "") for value in row_values):
                continue

            course_value = str(
                worksheet.cell(row=row_index, column=columns["Course"]).value or ""
            ).strip()
            if any(
                marker in course_value for marker in ("备注", "说明", "注：", "注:")
            ):
                note_rows.append(row_index)
                continue

            missing_fields = [
                field
                for field in required_fields
                if worksheet.cell(row=row_index, column=columns[field]).value
                in (None, "")
            ]
            if missing_fields:
                incomplete_rows.append(row_index)

        warnings: list[str] = []
        if note_rows:
            examples = "、".join(f"第 {row} 行" for row in note_rows[:3])
            suffix = "等" if len(note_rows) > 3 else ""
            warnings.append(
                f"CourseList 有 {len(note_rows)} 行疑似备注被写入课程明细："
                f"{examples}{suffix}"
            )
        if incomplete_rows:
            examples = "、".join(f"第 {row} 行" for row in incomplete_rows[:3])
            suffix = "等" if len(incomplete_rows) > 3 else ""
            warnings.append(
                f"CourseList 有 {len(incomplete_rows)} 行课程明细字段不完整："
                f"{examples}{suffix}"
            )
        return warnings

    @classmethod
    def _collect_sheet_quality_warnings(
        cls,
        worksheet,
        headers: list[Any],
        *,
        data_row_count: int,
        chart_count: int,
        conditional_formatting_count: int,
    ) -> list[str]:
        warnings: list[str] = []
        warnings.extend(
            cls._collect_dashboard_warnings(
                worksheet,
                data_row_count=data_row_count,
                chart_count=chart_count,
            )
        )
        warnings.extend(cls._collect_chart_warnings(worksheet))

        repeated_header_rows = cls._count_repeated_header_rows(worksheet, headers)
        if repeated_header_rows:
            warnings.append(
                f"{worksheet.title} 有 {repeated_header_rows} 行疑似重复表头"
            )

        warnings.extend(cls._collect_cleaned_data_warnings(worksheet, headers))

        if cls._sheet_expects_clean_text(worksheet.title):
            dirty_text_cells = cls._count_cells_with_line_breaks_or_tabs(worksheet)
            if dirty_text_cells:
                warnings.append(
                    f"{worksheet.title} 有 {dirty_text_cells} 个单元格仍包含换行或 Tab"
                )

        warnings.extend(
            cls._collect_gross_margin_warnings(
                worksheet,
                headers,
                conditional_formatting_count=conditional_formatting_count,
            )
        )
        warnings.extend(cls._collect_product_detail_warnings(worksheet, headers))
        warnings.extend(cls._collect_inventory_impact_warnings(worksheet, headers))
        warnings.extend(cls._collect_summary_sheet_warnings(worksheet, headers))
        warnings.extend(cls._collect_course_list_warnings(worksheet, headers))
        return warnings

    @classmethod
    def _collect_input_sheet_names(cls, input_paths: Sequence[Path] | None) -> set[str]:
        if not input_paths:
            return set()
        try:
            from openpyxl import load_workbook
        except ImportError:
            return set()

        sheet_names: set[str] = set()
        for input_path in input_paths:
            if input_path.suffix.lower() not in cls._EXCEL_FORMULA_SUFFIXES:
                continue
            workbook = None
            try:
                workbook = load_workbook(
                    input_path,
                    data_only=False,
                    read_only=True,
                )
                sheet_names.update(workbook.sheetnames)
            except Exception as exc:
                logger.debug(f"[文件管理] 读取输入工作簿 Sheet 名失败: {exc}")
            finally:
                if workbook is not None:
                    workbook.close()
        return sheet_names

    @classmethod
    def _build_excel_quality_summary(
        cls,
        output_path: Path,
        *,
        quality_warning_input_paths: Sequence[Path] | None = None,
    ) -> dict[str, Any] | None:
        if output_path.suffix.lower() not in cls._EXCEL_FORMULA_SUFFIXES:
            return None

        try:
            from openpyxl import load_workbook
        except ImportError:
            return None

        try:
            workbook = load_workbook(output_path, data_only=False, read_only=False)
        except Exception as exc:
            return {"file_type": "excel", "warnings": [f"无法读取质量摘要：{exc}"]}

        try:
            input_sheet_names = cls._collect_input_sheet_names(
                quality_warning_input_paths
            )
            sheet_summaries: list[dict[str, Any]] = []
            formula_count = 0
            chart_count = 0
            conditional_formatting_count = 0
            warnings: list[str] = []
            issues_rows: int | None = None
            issue_type_counts: dict[str, int] = {}

            for worksheet in workbook.worksheets:
                headers = [
                    worksheet.cell(row=1, column=column).value
                    for column in range(1, worksheet.max_column + 1)
                ]
                data_row_count = max(worksheet.max_row - 1, 0)
                sheet_chart_count = len(getattr(worksheet, "_charts", []))
                sheet_conditional_formatting_count = (
                    cls._count_conditional_formatting_rules(worksheet)
                )
                chart_count += sheet_chart_count
                conditional_formatting_count += sheet_conditional_formatting_count
                sheet_summaries.append(
                    cls._build_sheet_quality_summary(
                        worksheet,
                        data_row_count=data_row_count,
                        chart_count=sheet_chart_count,
                        conditional_formatting_count=(
                            sheet_conditional_formatting_count
                        ),
                    )
                )
                formula_count += cls._count_sheet_formulas(worksheet)
                if worksheet.title not in input_sheet_names:
                    warnings.extend(
                        cls._collect_sheet_quality_warnings(
                            worksheet,
                            headers,
                            data_row_count=data_row_count,
                            chart_count=sheet_chart_count,
                            conditional_formatting_count=(
                                sheet_conditional_formatting_count
                            ),
                        )
                    )
                sheet_issues_rows, sheet_issue_type_counts, issue_warnings = (
                    cls._collect_issues_sheet_state(
                        worksheet,
                        headers,
                        data_row_count=data_row_count,
                    )
                )
                if sheet_issues_rows is not None:
                    issues_rows = sheet_issues_rows
                    issue_type_counts.update(sheet_issue_type_counts)
                    warnings.extend(issue_warnings)

            return {
                "file_type": "excel",
                "sheet_count": len(workbook.worksheets),
                "sheets": sheet_summaries,
                "formula_count": formula_count,
                "chart_count": chart_count,
                "conditional_formatting_count": conditional_formatting_count,
                "issues_rows": issues_rows,
                "issue_type_counts": issue_type_counts,
                "warnings": warnings,
            }
        finally:
            workbook.close()

    async def deliver_generated_file(
        self,
        event: AstrMessageEvent,
        output_path: Path | None,
        *,
        success_message: str | None = None,
        block_quality_warnings: bool = False,
        quality_warning_input_paths: Sequence[Path] | None = None,
    ) -> GeneratedFileDeliveryResult:
        if output_path is None or not output_path.exists():
            logger.info(
                "[文件管理] 生成文件不存在，跳过发送: %s",
                str(output_path) if output_path is not None else "<none>",
            )
            return GeneratedFileDeliveryResult(status="missing")

        file_size = output_path.stat().st_size
        max_size = self._workspace_service.get_max_file_size()
        if file_size > max_size:
            try:
                output_path.unlink(missing_ok=True)
            except Exception as exc:
                logger.warning(f"[文件管理] 删除超限生成文件失败: {exc}")
            return GeneratedFileDeliveryResult(
                status="oversized",
                file_size=file_size,
                max_size=max_size,
            )

        validation_errors = self._validate_excel_formula_risks(output_path)
        if validation_errors:
            return GeneratedFileDeliveryResult(
                status="invalid",
                file_size=file_size,
                max_size=max_size,
                validation_errors=validation_errors,
            )

        quality_summary = self._build_excel_quality_summary(
            output_path,
            quality_warning_input_paths=quality_warning_input_paths,
        )
        quality_warnings = self._quality_warnings_requiring_review(quality_summary)
        if block_quality_warnings and quality_warnings:
            return GeneratedFileDeliveryResult(
                status="review_required",
                file_size=file_size,
                max_size=max_size,
                validation_errors=quality_warnings,
                quality_summary=quality_summary,
            )

        if success_message is None:
            await self._delivery_service.send_file_with_preview(event, output_path)
        else:
            await self._delivery_service.send_file_with_preview(
                event,
                output_path,
                success_message,
            )
        return GeneratedFileDeliveryResult(
            status="sent",
            file_size=file_size,
            max_size=max_size,
            quality_summary=quality_summary,
        )
